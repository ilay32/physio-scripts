from Tkinter import *
import tkFileDialog

import serial
import time
from openpyxl import Workbook
global inputFile
global line
#global count_line
count_line=0;
ser = serial.Serial('COM3', 9600, timeout=5)  # , timeout=5)
message = "";

def PathGet():
    global inputFile
    inputFile=tkFileDialog.askdirectory()

def StartRecord():
    #global line
    global count_line
    #count_line=0;
    line=[]
    print(int(num_Input.get()))
    print("OK")
    #print(str(button_Stop['state']))
    ser.write(b'v');
    ser.write(str(num_Input.get()));
    DisplayArduino(label_4)
    #while str(button_Stop['state'])=='normal':
       # print ser.readline()

def DisplayArduino(label_4):

    global count_line
    print count_line
    global line
    line=[]
    def count():
        global line
        arduino_line=ser.readline();
        line.append(arduino_line)
        label_4.config(text=str(arduino_line))
        label_4.after(1000, count)
        #count_line = count_line + 1;
    count()


def StopRecord():
    global line
    #global count_line
    ser.write(b's');
    #button_Stop["state"] = DISABLED;

    #print(str(button_Stop['state']))
    print(line)

    arduinodata1 = line#ser.readlines()
    arduinodata1 = [i.split('\r\n', 1)[0] for i in arduinodata1]
    print(arduinodata1)
    data=arduinodata1;
    StartTime=float(data[0].split(': ', 2)[1])
    print StartTime
    walk_start= [[0] for i in range(int(num_Input.get()))]
    walk_end = [[0] for i in range(int(num_Input.get()))]
    Time_walkS=    [[0]  for i in range(int(num_Input.get()))]
    Time_walkE=    [[0]  for i in range(int(num_Input.get()))]
    Velocity_Walk = [[0] for i in range(int(num_Input.get()))]
    #print (data[1].split(' ', 22));
    book = Workbook()
    sheet = book.active
    sheet.cell(row= 1, column=1).value="Sonar 1 Dist [cm]";
    sheet.cell(row= 1, column=2).value= "Sonar 1 Time [msec]";
    sheet.cell(row=1, column=3).value="Sonar 2 Dist [cm]";
    sheet.cell(row= 1, column=4).value= "Sonar 2 Time [msec]";
    sheet.cell(row=1, column=5).value="Velocity [m/sec]";
    sheet.cell(row= 1, column=6).value= "Distance Between Sonars [m]";
    for j in range(1, int(num_Input.get())+1):
        walk_start[j-1] = float(data[j].split(' ', 22)[7])
        walk_end[j-1] = float(data[j].split(' ', 22)[16])
        Time_walkS[j-1] = float(data[j].split(' ', 22)[11])-StartTime
        Time_walkE[j - 1] = float(data[j].split(' ', 22)[20])-StartTime
        Velocity_Walk[j-1]=float(format(float(numWalk_Input.get())*1000/(Time_walkE[j - 1]-Time_walkS[j-1]), '.2f'))# meter/sec
        print((Time_walkE[j - 1]-Time_walkS[j-1]))
        print ('\n')
        sheet.cell(row=j+1, column=1).value = walk_start[j-1];
        sheet.cell(row=j + 1, column=2).value = Time_walkS[j - 1];
        sheet.cell(row=j + 1, column=3).value = walk_end[j - 1];
        sheet.cell(row=j + 1, column=4).value = Time_walkE[j - 1];
        sheet.cell(row=j + 1, column=5).value = Velocity_Walk[j - 1];

    sheet.cell(row=2,column=6).value=float(numWalk_Input.get());
    book.save(inputFile + '\\' + str(text_Input.get()) + '.xlsx')


    print "walk dist\n"
    print walk_start
    print "time\n "
    print Time_walkS
    print "Velocity\n"
    print Velocity_Walk
    #ser.close()
    root.quit()



    #ser.write(b'v');


root=Tk();
root.geometry("600x600+0+0")
root.title("calculate distance")
text_Input=StringVar()
num_Input=IntVar()
numWalk_Input=DoubleVar()
operator = ""
Tops=Frame(root,width=700,height = 50 , bg="powder blue", relief=SUNKEN)
Tops.pack(side=TOP)

Lb=Label(Tops, font=('ariel',30,'bold'), text= "Distance Calculation:",fg="Steel Blue",bd=10,anchor='w')
Lb.grid(row=0,column=0)
f1=Frame(root,width=800,height = 700 , bg="powder blue", relief=SUNKEN)
f1.pack(side=TOP)
button_Path=Button(f1,text="Choose Path To Save File",bg="powder blue",command = PathGet,font="-weight bold")
button_Path.grid(columnspan=2)
label_1=Label(f1,font=('ariel',20,'bold'),text="Name of The Subject",bg="powder blue")
label_1.grid(columnspan=2)
txtDisp1=Entry(f1,font=('ariel',20,'bold'),textvariable=text_Input,bd=10,insertwidth=1,bg="powder blue",justify='left')
txtDisp1.grid(columnspan=2)
label_2=Label(f1,font=('ariel',20,'bold'),text="Number Of Walks",bg="powder blue")
label_2.grid(columnspan=2)
txtDisp2=Entry(f1,font=('ariel',20,'bold'),textvariable=num_Input,bd=10,insertwidth=1,bg="powder blue",justify='left')
txtDisp2.grid(columnspan=2)
label_3=Label(f1,font=('ariel',20,'bold'),text="Distance of Walks [M] ",bg="powder blue")
label_3.grid(columnspan=2)
txtDisp3=Entry(f1,font=('ariel',20,'bold'),textvariable=numWalk_Input,bd=10,insertwidth=1,bg="powder blue",justify='left')
txtDisp3.grid(columnspan=2)
button_Start=Button(f1,text="START RECORD",bg="green",command = StartRecord,font="-weight bold")
button_Start.grid(columnspan=2)
button_Stop=Button(f1,text="STOP RECORD",bg="red",command = StopRecord,font="-weight bold")
button_Stop.grid(columnspan=2)
label_4=Label(f1,font=('ariel',20,'bold'),text="Arduino Data",bg="powder blue")
label_4.grid(columnspan=2)
#label_4.pack()
#DisplayArduino(label_4)
root.mainloop()


