import webbrowser,threading,time,json,os,re,glob,platform,math,yaml
import http.server as hs
from openpyxl import Workbook,load_workbook

HOST_NAME = 'localhost'
PORT_NUMBER = 8000

class ExpRunner(hs.SimpleHTTPRequestHandler):
    #savedir = os.path.join(os.path.abspath(os.pardir),"corridor-walks")
    #savedir = "corridor-walks"
    savedir = os.path.join(os.path.expanduser("~") ,"Desktop","corridor-walks")
    datacols = [
        "walk no.", 
        "absolute start", 
        "absolute end", 
        "relative start", 
        "relative end", 
        "duration", 
        "speed m/s", 
        "distractor digits",
        "remembered as"
    ]
    
    def do_HEAD(self):
        self._set_headers()

    def _set_headers(self):
        self.send_response(200)
        self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        self.send_header("Content-type","application/json")
        self.end_headers()

    def respond(self,string):
        self.wfile.write(bytes(string,"utf-8"))
    
    def do_GET(self):
        m = re.match(r'(.*download\/\?file\=)(.+\.[a-z]{2,4})$',self.path)
        if m and len(m.groups()) == 2:
            self.path = os.path.join(ExpRunner.savedir,m.group(2))
        return hs.SimpleHTTPRequestHandler.do_GET(self)

    def do_POST(self):
        post = json.loads(self.rfile.read(int(self.headers['Content-Length'])).decode("utf-8"))
        command = post['command']
        self._set_headers()
        if command == "stopper_validation":
            try:
                tfile = 'wcapp/stopper-validation.png'
                import numpy as np
                import matplotlib.pyplot as plt
                x = np.array(list(post['data'].keys())).astype(float)
                y = np.mean(list(post['data'].values()),axis=1)
                A = np.vstack([x, np.ones(len(x))]).T
                m,c = np.linalg.lstsq(A,y,rcond=-1)[0]
                plt.plot(x, y, 'o', label='Original data', markersize=10)
                plt.plot(x, m*x + c, 'r', label='Fitted line')
                plt.legend()
                plt.grid()
                os.remove(tfile)
                plt.savefig(tfile)
                #plt.show()
                self.respond(json.dumps({"response":{
                        "a" : m,
                        "b": c
                }}))
            except Exception as e:
                self.respond(json.dumps({
                        "response": str(e)
                }))

                                     
        if command == "get_experiment":
            with open(os.path.join(os.path.expanduser("~") ,"Desktop","experiment-sequence.csv")) as exp:
                experiment = list()
                try:
                    for l in exp.readlines():
                        l = l.strip()
                        s = list(map(str.strip,l.split(",")))
                        if s[0] == "title":
                            continue
                        assert re.match(r'^[\w\s_–\-]+,\d+,[A-Z]',l),"invalid csv line: {}".format(l)
                        experiment.append(dict(
                            title=s[0],
                            duration=int(s[1]),
                            stype=s[2]
                        ))
                    self.respond(json.dumps({
                        "response" : "parsed",
                        "experiment" : experiment
                    }))

                except Exception as e:
                    self.respond(json.dumps({
                        "response": "could not load the sequence:\n{}".format(str(e))
                    }))

        if command  == "check_existing":
            dest = os.path.join(ExpRunner.savedir,post['filename'])
            wb = load_workbook(dest)
            attempted = post['savesheet'] 
            other = 'restep' if attempted == 'normal' else 'normal'
            status = "exists" if attempted in wb.sheetnames else "absent"
            if status == "exists" and other in wb.sheetnames and attempted != "single":
                status = "both_exist"
            firstcond = self.find_in_kvrow(wb,"first condition") 
            num,pause = "",""
            if post['savesheet'] == "single":
                try:
                    num,pause = self.compute_singletask_params(wb)
                except Exception as e:
                    self.respond(json.dumps({"error": str(e)}))
                    return
            self.respond(json.dumps({
                "sheet_status" : status,
                "first" : firstcond,
                "numtrials" : num,
                "pausetime" : pause,
                "nirs41" : self.find_in_kvrow(wb,"nirs41") 
            }))
        if command  == 'checkfile':
            f,e = os.path.splitext(post['filename'])
            if os.path.isfile(os.path.join(ExpRunner.savedir,f+e)):
                f = self.findlast(f)
            self.respond(json.dumps({"please_saveto": f+e}))
        
        if command == 'stop':
            print("shutting down server")
            server_stop.start()
            self.respond(json.dumps({"response" : "the server has shutdown"}))
        
        if command == 'save':
            try:
                if post['savesheet'] == 'single':
                    self.save_single_data(post)
                else:
                    self.save_walk_data(post)
                self.respond(json.dumps({"response" : "saved"}))
                print("saved to "+post['savefile'])
            except Exception as e:
                self.respond(json.dumps({"response" : "failed" , "error" : str(e)}))
                print(str(e))
        #return SimpleHTTPServer.SimpleHTTPRequestHandler.do_POST(self)
    
    def find_in_kvrow(self,wrkbk,field):
        for s in wrkbk.sheetnames:
            for row in wrkbk[s].iter_rows():
                if row[0].value == field:
                    return row[1].value
        return

    def findlast(self,file_base):
        f = file_base
        fs = glob.glob(os.path.join(ExpRunner.savedir,f+"([0-9])*"))
        if len(fs) == 0:
            return f+"(1)"
        latest = 1
        for dup in fs:
            m = re.match(r'^(.*)(\(\d+\))(\.\w{2,4})$',dup)
            if len(m.groups()) == 3:
                i = int(m.group(2).strip("()"))
                if i > latest:
                    latest = i
        return  f+"("+str(latest+1)+")"
    
    def compute_singletask_params(self,wb):
        if "restep" not in wb.sheetnames or "normal" not in wb.sheetnames:
            raise(Exception("the selected file must contain walks data"))
            return
        durations = list()
        cols = ExpRunner.datacols
        for sheet in ['normal','restep']:
            for r in wb[sheet].iter_rows():
                dur = r[cols.index('duration')].value
                dig = r[cols.index('distractor digits')].value
                if dig and str.isdigit(dig):
                    durations.append(dur*1000)
        if(len(durations) > 0):
            n,a = math.floor(len(durations)/2),float(sum(durations))/len(durations)
            print("single: {:d} trials with {:.3f} pausetime".format(n,a))
            return n,a      
        else:
            raise(Exception("could not extract data from file"))
    
    def save_single_data(self,p):
        dest = os.path.join(ExpRunner.savedir,p['savefile'])
        sheetname = 'single'
        wb = load_workbook(dest)
        g = p['data']['globals'] 
        start = g['start_time']
        if sheetname in wb.sheetnames:
            s = wb[sheetname]
            wb.remove(s)
        datsheet = wb.create_sheet(sheetname)
        
        # header row
        datsheet.append(["trial no.", "retention start","sequence","answer"])
        # data
        for i,t in enumerate(p['data']['trialdata'],1):
            s,r,a = t # sequence retention (starting point) answer
            row = [i,float(r - start)/1000,s,a]
            datsheet.append(row)
        
        # params
        datsheet.append([""])
        datsheet.append(["researcher",g['walks']['researcher']])
        datsheet.append(["pausetime",float(g['single']['pausetime'])/1000])
        self.comment_and_start(g,datsheet)
        wb.save(filename = dest)

    def comment_and_start(self,globj,wbsheet):
        # write the start time 
        start = globj['start_time']
        abstart = float(start)/1000
        wbsheet.append([
            "start:",
            time.ctime(abstart),
            "unix stamp:",
            abstart
        ])

        # write the comment if there is one 
        if globj.get('comments') is not None:
            wbsheet.append(["preliminary comments",globj['comments'].replace("\n"," ")])
        if globj.get('postcomments') is not None:
            wbsheet.append(["post hoc comments",globj['postcomments'].replace("\n"," ")])
 
    def save_walk_data(self,p):
        dat = p['data']
        if not os.path.isdir(ExpRunner.savedir):
            os.mkdir(ExpRunner.savedir)
        dest = os.path.join(ExpRunner.savedir,p['savefile'])
        g = dat['globals']
        subj = g['subject']
        walks = g['walks']
        start = g['start_time']
        sheetname = p['savesheet']
        if g['newSubject']: 
            wb = Workbook()
            datsheet = wb.active
            datsheet.title = sheetname
        else:
            wb = load_workbook(dest)
            if sheetname in wb.sheetnames:
                s = wb[sheetname]
                wb.remove(s)
            datsheet = wb.create_sheet(sheetname)
        
        # header row
        for i,c in enumerate(ExpRunner.datacols,1):
            datsheet.cell(column=i,row=1,value=c)
        
        # data
        for i,w in enumerate(dat['walkdata']):
            isdist = str(i) in dat['distractions'].keys()
            s,e = w
            row = [i+1,s,e,float(s - start)/1000, float(e - start)/1000,float(e - s)/1000,1000*float(walks['distance'])/(e - s)]
            if isdist: 
                row += dat['distractions'][str(i)]
            else:
                row += ["",""]
            datsheet.append(row)

        # write the parameters in the same sheet
        for i,(k,v) in enumerate(walks.items(),len(dat['walkdata'])+3):
            datsheet.cell(row=i,column=1,value=k.replace("_"," "))
            datsheet.cell(row=i,column=2,value=v)
        
        self.comment_and_start(g,datsheet)

        # if not adding to existing file, write the subject details 
        if g['newSubject']:
            subsheet = wb.create_sheet(title="subject")
            for i,(k,v) in enumerate(subj.items(),1):
                if subj['regular_sport'] == "none" and k == 'weekly_minutes':
                    continue
                subsheet.cell(row=i,column=1,value=k.replace("_"," "))
                subsheet.cell(row=i,column=2,value=v)
            # new subject -- specify their nirs41 location
            subsheet.append(['nirs41',g['nirs41']]) 
            subsheet.append(['first condition',walks['shoe_type']])
            # and finally save as "dest"
        wb.save(filename = dest)


httpd = hs.HTTPServer((HOST_NAME, PORT_NUMBER), ExpRunner)
    
def runserver():
    print(time.asctime(), "Server Starts - %s:%s" % (HOST_NAME, PORT_NUMBER))
    httpd.serve_forever()
    

def killserver():
    httpd.shutdown()
    httpd.socket.close()


if __name__ == '__main__':
    if platform.system() == "Windows":
        os.chdir(os.path.dirname(__file__))
    #server_go = threading.Thread(None,runserver)
    #server_go.start()
    server_stop = threading.Thread(None,killserver)
    server_stop.setDaemon(True)
    webbrowser.open('http://localhost:8000/wcapp/')
    runserver()

